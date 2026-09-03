#!/usr/bin/env python3
"""Веса нетто и габариты упаковки из юнит-экономики владелицы.

Источник — tasks/excel-from-owner/UNIT_economics_cogs_tariffs.xlsx, лист «Юнит».
Чтение колонок подтверждено владелицей 03.09.2026:
    Вес                     — вес НЕТТО, чистый, без упаковки
    Длина · Ширина · Высота — габариты УПАКОВКИ, не предмета
    Упаковка                — тип пакета: «своя», «ОРР 24*34», «слайдер 25*30»
Габаритов предмета в файле нет — для них нужен замер.

Запуск:
    python3 scripts/export_unit_weights.py [--out docs/seo/unit-weights.json]
"""
import argparse
import re
import json
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "tasks" / "excel-from-owner" / "UNIT_economics_cogs_tariffs.xlsx"

# Колонки листа «Юнит». Шапка занимает три строки, данные с четвёртой.
COL_ART, COL_NAME = 3, 4
COL_LEN, COL_WID, COL_HEI, COL_NET, COL_PACK, COL_BOX = 12, 13, 14, 15, 16, 17


ARTICLE_RE = re.compile(r"[A-Z][A-Z0-9]{7,}")


def num(v):
    """Число или None. Ноль в файле значит «не заполнено», не «нулевой вес»."""
    if isinstance(v, (int, float)) and v:
        return v
    return None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--out", default=str(ROOT / "docs" / "seo" / "unit-weights.json"))
    args = ap.parse_args()

    # read_only=True занижает max_column — намеренно не используем (урок 03.09).
    ws = openpyxl.load_workbook(SRC, data_only=True)["Юнит"]

    rows = {}
    for r in range(4, ws.max_row + 1):
        art = ws.cell(r, COL_ART).value
        if not art or not isinstance(art, str):
            continue
        art = art.strip()
        # В листе «Юнит» под шапкой идёт строка нумерации колонок: в графе
        # артикула там оказалась «3», и она попадала в выгрузку пустой
        # записью, а оттуда — строкой-призраком в лист «Размеры и вес».
        # Артикулы АРОЛС — заглавная латиница с цифрами, не короче восьми.
        if not ARTICLE_RE.fullmatch(art):
            continue
        net_kg = num(ws.cell(r, COL_NET).value)
        rec = {
            "name": str(ws.cell(r, COL_NAME).value or "").strip(),
            "net_g": round(net_kg * 1000) if net_kg else None,
            "pkg_len": num(ws.cell(r, COL_LEN).value),
            "pkg_wid": num(ws.cell(r, COL_WID).value),
            "pkg_hei": num(ws.cell(r, COL_HEI).value),
            "pkg_type": (str(ws.cell(r, COL_PACK).value).strip()
                         if ws.cell(r, COL_PACK).value is not None else None),
            "in_box": num(ws.cell(r, COL_BOX).value),
        }
        rows[art] = rec

    out = {
        "_": "Вес нетто и габариты упаковки. Источник — юнит-экономика владелицы, "
             "лист «Юнит». Чтение колонок подтверждено владелицей 03.09.2026: "
             "Вес — нетто, Длина/Ширина/Высота — упаковка. Габаритов предмета в файле нет. "
             "Пересобрать: python3 scripts/export_unit_weights.py",
        "articles": rows,
    }
    Path(args.out).write_text(
        json.dumps(out, ensure_ascii=False, indent=1), encoding="utf-8")

    with_net = sum(1 for v in rows.values() if v["net_g"])
    print(f"артикулов: {len(rows)}, с весом нетто: {with_net} → {args.out}")


if __name__ == "__main__":
    main()
