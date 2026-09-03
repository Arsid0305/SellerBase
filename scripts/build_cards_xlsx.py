#!/usr/bin/env python3
"""Сборка книги для заливки карточек в кабинет WB из docs/seo/descriptions/*.md.

Один лист на группу. Колонки — то, что вносится руками: наименование, значения
полей и текст описания. Значения полей разделены переводом строки внутри ячейки:
в кабинете каждое значение вносится отдельной строкой (склейка через запятую —
дефект заполнения, см. tasks/rules.md).

    python3 scripts/build_cards_xlsx.py [--out kartochki_arols.xlsx]

Файлы описаний написаны в разное время и в разной разметке, поэтому у каждого
свой разборщик. Общий контракт для новых файлов — раздел на артикул:

    ## ACRA7TB201BC · 186826338
    | поле | значение |
    | **Наименование** | `...` |
    ### Описание — N знаков
    ```
    текст
    ```
"""
import argparse, json, pathlib, re, sys

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError:
    sys.exit("нужен openpyxl: pip install openpyxl")

ROOT = pathlib.Path(__file__).resolve().parent.parent
DESCR = ROOT / "docs" / "seo" / "descriptions"

ART = r"[A-Z]{4}\d[A-Z]{2}\d{3}[A-ZС]{2}"  # «С» в конце части артикулов кириллическая


def fenced(text):
    """Первый блок в тройных кавычках."""
    m = re.search(r"```\n(.*?)\n```", text, re.S)
    return m.group(1).strip() if m else ""


def field(section, name):
    """Значение из строки таблицы «| поле | значение |»."""
    m = re.search(r"^\|\s*\**" + re.escape(name) + r"\**\s*\|\s*(.+?)\s*\|\s*$",
                  section, re.M)
    if not m:
        return ""
    v = m.group(1).replace("`", "").strip()
    return "" if v.startswith("стоит") else v.replace(" · ", "\n")


def split_sections(text, level="## "):
    """Разбить на разделы по заголовку, вернуть [(заголовок, тело)]."""
    parts = re.split(r"(?m)^" + re.escape(level) + r"(.+)$", text)
    return list(zip(parts[1::2], parts[2::2]))


# Поля, которые в лист не выносим: они служебные либо дублируют другие колонки.
SKIP_FIELDS = {
    "поле", "Описание",
    # Служебные поля карточки: в книгу для заливки не идут.
    "Артикул OZON", "Баркод", "NTIN", "ИКПУ", "Код ТРУ 1", "Код ТРУ 2",
    "Ставка НДС", "Тип доставки", "Код упаковки",
    # Разбор массажёров записывает габариты слитно — «Высота / Ширина / Глубина
    # предмета | 3 · 7 · 110». Для книги они уже разложены по отдельным колонкам
    # из packaging.json, а слитная строка дала бы дубль колонки.
    "Вес товара с упаковкой",
}


def is_skipped(name):
    return name in SKIP_FIELDS or " / " in name

# Три состояния поля, и в книге они должны различаться на глаз.
# Раньше «проверено, верно» и «до поля не дошли» выглядели одинаково — пустой
# ячейкой, и при заливке было не понять, это «не трогать» или «не заполнено».
OK = "верно, не трогать"
MEASURE = "нужен замер"


def status(value):
    """Значение поля для книги. Пусто — значит поле не разобрано.

    «30 — стоит верно» превращается в «30 — верно, не трогать»: видно и что стоит
    в карточке, и что менять не нужно. Голое «стоит верно» — только пометку.
    """
    v = value.replace("`", "").replace("**", "").strip()
    if not v or v.startswith("---"):
        return ""
    low = v.lower()
    if low.startswith("требует") or low.startswith("замерить"):
        return MEASURE
    if low.startswith(("стоит верно", "оставить", "стоит", "то же")):
        return OK
    for marker in (" — стоит верно", " - стоит верно", ", стоит верно"):
        if marker in v:
            return v.replace(marker, "").strip() + f" — {OK}"
    if "требует замера" in low:
        return MEASURE
    return v.replace(" · ", "\n")


def all_fields(section):
    """Все строки таблицы «| поле | значение |» раздела, как есть.

    Раньше здесь был список из четырёх имён, и любое разобранное поле сверх него
    молча терялось: «Упаковка» стояла во всех десяти таблетницах и ни разу
    не попала в книгу. Теперь колонки набираются из того, что реально разобрано.
    """
    out = {}
    for name, value in re.findall(r"^\|\s*\**([^|*]+?)\**\s*\|\s*(.+?)\s*\|\s*$",
                                  section, re.M):
        name = name.strip()
        if is_skipped(name) or name.startswith("-"):
            continue
        v = status(value)
        if v:
            out[name] = v
    return out


def parse_generic(text):
    """Контракт «раздел на артикул»: ## ART · nm_id."""
    rows = []
    for head, body in split_sections(text):
        m = re.match(r"\s*(" + ART + r")\s*·\s*(\d+)", head)
        if not m:
            continue
        row = {"Артикул": m.group(1), "nm_id": m.group(2), "Описание": fenced(body)}
        row.update(all_fields(body))
        rows.append(row)
    return rows


def parse_roller(text):
    """massagers-roller-101.md: ## RJ · оранжевый · 223420736 + общая таблица."""
    common = text.split("## Общее для всех трёх", 1)[-1].split("\n---", 1)[0]
    rows = []
    for head, body in split_sections(text):
        m = re.match(r"\s*([A-Z]{2})\s*·\s*([^·]+?)\s*·\s*(\d+)", head)
        if not m:
            continue
        # Общая таблица разбирается целиком, а не по списку имён: раньше сюда
        # были вписаны четыре поля, и «Упаковка» из неё в книгу не попадала,
        # хотя стояла в разборе. Та же ошибка, что была у таблетниц с этим полем.
        row = all_fields(common)
        row.update({k: v for k, v in all_fields(body).items() if v})
        row.update({
            "Артикул": "ACRB1MS101" + m.group(1),
            "nm_id": m.group(3),
            "Наименование": field(body, "Наименование"),
            "Зона массажа": field(body, "Зона массажа"),
            "Цвет": field(body, "Цвет"),
            "Описание": fenced(body),
        })
        rows.append({k: v for k, v in row.items() if v})
    return rows


def parse_balls(text):
    """myachi-106-107-109.md: наименования в общей таблице, описания в ### <код>."""
    titles, zones = {}, {}
    for code, title in re.findall(r"\|\s*`(" + ART + r")`\s*\|\s*`([^`]+)`", text):
        titles[code] = title
    for d, vals in re.findall(r"\|\s*(\d)\s*см\s*\|\s*((?:`[^`]+`\s*·?\s*)+)\|", text):
        zones["10" + d] = "\n".join(re.findall(r"`([^`]+)`", vals))
    action = "\n".join(["снятие мышечного напряжения", "снятие отечности",
                        "улучшение кровообращения"])

    # Цвет разобран одной таблицей «| SKU | стоит сейчас | ставим |» в разделе
    # «### Цвет». Ищем строго внутри него: снаружи есть другие таблицы с кодом
    # в первой колонке, и без границы раздела в «Цвет» попадали их числа.
    # «оставить» значит «текущее значение верно». Раньше такие в книгу не выносили:
    # пустая ячейка означала «не трогать». С 02.09 контракт обратный — пусто значит
    # «не разобрано», — поэтому status() превращает «оставить» в «верно, не трогать»,
    # и восемь мячей перестали показывать пустой «Цвет» там, где он проверен.
    colors = {}
    csec = re.search(r"(?m)^### Цвет\b.*?(?=^#{2,3} |\Z)", text, re.S)
    if csec:
        for line in csec.group(0).split("\n"):
            cells = [c.strip() for c in line.strip().strip("|").split("|")]
            if len(cells) != 3 or not cells[0].startswith("`"):
                continue
            put = status(cells[2].replace(" + ", "\n"))
            if not put:
                continue
            for code in re.findall(r"`(\d{3}[A-ZС]{1,2})`", cells[0]):
                colors["ACRB1MS" + code] = put

    rows = []
    for head, body in split_sections(text, "### "):
        m = re.match(r"\s*(\d{3}[A-ZС]{1,2})\s*·", head)
        if not m:
            continue
        code = "ACRB1MS" + m.group(1)
        if code not in titles:
            continue
        rows.append({
            "Артикул": code,
            "nm_id": "",
            "Наименование": titles[code],
            "Зона массажа": zones.get(code[7:10], ""),
            "Действие": action,
            "Материал изделия": "ПВХ",
            "Цвет": colors.get(code, ""),
            # Одно значение на всю группу: в поставке только сам мяч.
            # Плёнка ОПП — материал упаковки, в значение поля не идёт:
            # покупатель так не ищет, а «Упаковка» отвечает на другой вопрос.
            "Комплектация": "массажный мяч",
            "Упаковка": "пакет с клапаном",
            "Описание": fenced(body),
        })
    return rows


INFO = ROOT / "docs" / "seo" / "infographics"

def parse_bags(text):
    """meshki-dlya-stirki.md: поля таблицей «колонка на SKU», описания в «### ART · имя».

    Разбор идёт по подгруппам: внутри каждой одна таблица «Значения полей», где
    первая колонка — имя поля, остальные — артикулы. Готовы не все подгруппы,
    поэтому карточка попадает в лист только если у неё есть описание.
    """
    rows = []
    for _, body in split_sections(text):
        fields = {}
        m = re.search(r"^\| поле \|(.+?)\|\s*$", body, re.M)
        if m:
            skus = [re.sub(r"[`\s]|\(.*?\)", "", c) for c in m.group(1).split("|") if c.strip()]
            start = body.index(m.group(0))
            for line in body[start:].split("\n")[2:]:
                if not line.startswith("|"):
                    break
                cells = [c.strip() for c in line.strip("|").split("|")]
                # Имя поля бывает выделено жирным: «**Высота предмета**» и
                # «Высота предмета» — одно поле, но без чистки станут двумя колонками.
                cells[0] = cells[0].replace("*", "").replace("`", "").strip()
                for sku, val in zip(skus, cells[1:]):
                    v = status(val.replace(" + ", "\n"))
                    if v:
                        fields.setdefault(sku, {})[cells[0]] = v

        # «**`SKU`** — «Комплектация», N значений:» + блок в тройных кавычках
        for m2 in re.finditer(r"\*\*`(\d[A-Z]{2}\d{3}[A-ZС]{2})`\*\* — «Комплектация»"
                              r"(?:(?!\*\*`)[\s\S])*?```\n(.*?)\n```",
                              body, re.S):
            fields.setdefault(m2.group(1), {})["Комплектация"] = m2.group(2).strip()

        for head, sec in split_sections(body, "### "):
            mm = re.match(r"\s*`?(" + ART + r")`?\s*·\s*(.+)", head)
            if not mm:
                continue
            sku = mm.group(1)
            row = {
                "Артикул": sku,
                "Наименование": mm.group(2).strip(),
                "Описание": fenced(sec),
            }
            # Все разобранные поля подгруппы, а не два выбранных: у мешков
            # набор полей предмета свой, и что именно разобрано — решает файл.
            row.update({k: v for k, v in fields.get(sku[4:], {}).items()
                        if k not in SKIP_FIELDS})
            rows.append(row)
    return rows


INFO_GROUPS = [
    ("Мешки для стирки", "meshki-stirki.md"),
    ("Мячи 6-7-9 см", "myachi-106-107-109.md"),
    ("Массажёры 101", "massagers-roller-101.md"),
    ("Таблетницы", "tabletnitsy.md"),
]


def parse_order(text):
    """Нумерованные списки из разделов «Порядок».

    В файле их может быть несколько — по одному на подгруппу, и заголовок
    бывает как второго уровня, так и третьего. Берём все и склеиваем.
    """
    items = []
    for m in re.finditer(r"(?m)^#{2,3} Порядок\s*$(.*?)(?=^#{2,3} |\Z)", text, re.S):
        for i in re.findall(r"(?m)^\d+\.\s+(.+?)(?=\n\d+\.|\n\n|\Z)", m.group(1), re.S):
            items.append(re.sub(r"\s+", " ", i).replace("**", "").strip())
    return items


def parse_per_article(text):
    """Таблицы «По артикулам»: артикул -> что сделать со слайдами.

    Разделов может быть несколько (по подгруппам), уровень заголовка ## или ###.
    """
    out = {}
    for m in re.finditer(r"(?m)^#{2,3} По артикулам\s*$(.*?)(?=^#{2,3} |\Z)", text, re.S):
        for art, todo in re.findall(r"(?m)^\|\s*(" + ART + r")\s*\|\s*(.+?)\s*\|\s*$", m.group(1)):
            out[art] = todo.strip()
    return out


def all_per_article():
    """Задания по слайдам со всех групп разом."""
    out = {}
    for _, fname in INFO_GROUPS:
        path = INFO / fname
        if path.exists():
            out.update(parse_per_article(path.read_text(encoding="utf-8")))
    return out


def urgency(text):
    """Срочность — по наличию медицинских заявлений в задании."""
    return "не срочно" if "нет ни одного медицинского заявления" in text else "срочно"


def sheet_info(wb):
    rows = []
    for name, fname in INFO_GROUPS:
        path = INFO / fname
        if not path.exists():
            continue
        t = path.read_text(encoding="utf-8")
        for i, item in enumerate(parse_order(t), 1):
            # Группу и срочность повторяем в каждой строке: лист фильтруют и
            # сортируют, и при протяжке только в первой строке пункты теряют группу.
            rows.append([name, urgency(t), i, item])
    if not rows:
        return 0
    ws = wb.create_sheet("Инфографика")
    ws.append(["Группа", "Срочность", "№", "Что сделать"])
    for c in ws[1]:
        c.font = Font(bold=True)
    for r in rows:
        ws.append(r)
    for col, w in zip("ABCD", (20, 14, 5, 105)):
        ws.column_dimensions[col].width = w
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"
    return len(rows)


GROUPS = [
    ("Таблетницы", "tabletnitsy.md", parse_generic),
    ("Мячи 6-7-9 см", "myachi-106-107-109.md", parse_balls),
    ("Массажёры 101", "massagers-roller-101.md", parse_roller),
    ("Мешки для стирки", "meshki-dlya-stirki.md", parse_bags),
    ("Игрушки для животных", "igrushki-dlya-zhivotnyh.md", parse_generic),
    ("Скакалки", "skakalki.md", parse_generic),
    ("Фитнес-резинки", "espandery.md", parse_generic),
]

# Порядок колонок в листе. Список задаёт только очерёдность известных полей —
# всё, что разобрано сверх него, дописывается в конец само (см. sheet()).
# Хардкод списка стоил «Упаковки» у десяти таблетниц: поле было разобрано,
# но в книгу не попало, потому что имени не было в этом перечне.
COLS = ["Артикул", "nm_id", "Наименование", "Полное наименование товара",
        "Зона массажа", "Действие", "Материал изделия", "Цвет", "Комплектация",
        "Упаковка", "Код ТН ВЭД",
        "Высота предмета", "Ширина предмета", "Глубина предмета",
        "Вес товара без упаковки (г)",
        "Длина упаковки", "Ширина упаковки", "Высота упаковки",
        "Вес товара с упаковкой (г)",
        "Описание", "Рич-контент", "Инфографика",
        "Заказов за год", "Вопросов покупателей", "Остаток на складах",
        "Себестоимость, ₽"]

COUNTRY = "Китай"
BRAND = "АРОЛС"

# Габариты и вес в упаковке — docs/seo/packaging.json, выгрузка из sku_catalog.
# Это размеры В УПАКОВКЕ, а не изделия: цифры пришли из юнит-экономики, и разбор
# таблетниц (§«Размер и вес») прямо оговаривает, что годятся они для полей
# «Длина / Ширина / Высота упаковки» и логистики, но не для размеров предмета.
PACKAGING = json.loads((ROOT / "docs" / "seo" / "packaging.json").read_text(encoding="utf-8")) \
    if (ROOT / "docs" / "seo" / "packaging.json").exists() else {}

# Две разные пары, и путать их нельзя.
#
# Предмет: габариты и вес нетто — из карточек кабинета, там они уже стоят.
# Упаковка: габариты — из sku_catalog (юнит-экономика), они закономерно больше
# габаритов предмета: у таблетницы 201 предмет 15x9x4, упаковка 16x12x5.
#
# Вес брутто лежит в sku_catalog в dimensions.weightBrutto, и это подтверждено
# независимо: выгрузка кабинета «Общие характеристики одним файлом» от 02.09
# даёт те же цифры во всех 58 строках. Раньше колонка стояла пустой по всей
# книге — я забраковал соседнее поле unit_weight_kg (оно у шести таблетниц
# меньше нетто) и ошибочно заключил, что брутто нет вовсе. Поля разные.
#
# Конфликт «брутто меньше нетто» никуда не делся, но он не про источник:
# он есть и в самом кабинете у семи карточек — двух канатов и пяти таблетниц.
# Это дефект данных владелицы, и лист размеров теперь на него указывает
# отдельной строкой вместо прежнего сквозного «нет веса брутто».
# Факты каталога — docs/seo/catalog-facts.json: остаток, заказы за год, число
# вопросов покупателей, себестоимость. В карточку они не заливаются: это
# справочные колонки, чтобы при заливке было видно, насколько карточка важна
# и есть ли по ней вопросы, которые стоит перечитать. Стоят последними,
# после всех заливочных полей.
FACTS = json.loads((ROOT / "docs" / "seo" / "catalog-facts.json").read_text(encoding="utf-8")) \
    if (ROOT / "docs" / "seo" / "catalog-facts.json").exists() else {}

# Снимок кабинета для групп, по которым разбора ещё НЕ было: владелица 03.09
# попросила листы, но разбор отложила до среза Эвирмы. Поэтому здесь не
# «что вносить», а «что стоит сейчас» — колонки так и подписаны, чтобы лист
# нельзя было спутать с заданием на заливку.
SNAPSHOT = json.loads((ROOT / "docs" / "seo" / "cabinet-snapshot.json").read_text(encoding="utf-8")) \
    if (ROOT / "docs" / "seo" / "cabinet-snapshot.json").exists() else {}

NO_ANALYSIS = "разбора не было — ждём срез Эвирмы"

FACT_FIELDS = [
    ("Заказов за год", "orders_year"),
    ("Вопросов покупателей", "questions"),
    ("Остаток на складах", "stock"),
    ("Себестоимость, ₽", "cost_rub"),
]


def fill_facts(row):
    """Справочные цифры по артикулу плюс nm_id, если его не было в разборе."""
    f = FACTS.get(row.get("Артикул", ""))
    if not f:
        return
    if f.get("nm_id"):
        row.setdefault("nm_id", f["nm_id"])
    for col, key in FACT_FIELDS:
        if f.get(key) is not None:
            row.setdefault(col, f[key])


PACK_FIELDS = [
    ("Высота предмета", "item_hei"),
    ("Ширина предмета", "item_wid"),
    ("Глубина предмета", "item_dep"),
    ("Вес товара без упаковки (г)", "net_g"),
    ("Длина упаковки", "len"),
    ("Ширина упаковки", "wid"),
    ("Высота упаковки", "hei"),
    ("Вес товара с упаковкой (г)", "brutto_g"),
]


def fill_packaging(row):
    """Габариты и вес по артикулу. Пусто — значит в каталоге нет цифры."""
    pack = PACKAGING.get(row.get("Артикул", ""))
    if not pack:
        return
    for col, key in PACK_FIELDS:
        v = pack.get(key)
        if v is not None:
            row.setdefault(col, v)

# Коды ТН ВЭД по префиксу артикула. Источник — docs/marking/codes-marking.md,
# который, в свою очередь, собран по матрице владелицы (выбор помечен зелёным).
# Ключ проверяется как префикс, длинные ключи имеют приоритет над короткими.
TNVED = {
    "ACRA7TB": "3926909709",      # таблетницы
    "AHMA": "6307909800",         # мешки для стирки
    "ACRB1MS": "9506919000",      # мячи массажные
    "ACRH1MS": "9506919000",      # массажёр для пальцев
    "ACRF1BN101": "6307909800",   # бандажи лайкра
    "ACRF1BN201": "3926909709",   # бандаж силиконовый
    "ACRB1SK": "9506919000",      # скакалки
    "ACRB3FR": "9506919000",      # эспандеры
    "ACRB4FR": "9506919000",
    "ACRB5FR": "9506919000",
    "AKTA4ST": "3923509000",      # пробки для бутылок
    "AKTA2KN101": "3926909709",   # шпатель кондитерский — см. вопрос 2 в codes-marking.md
    "AKTA2KN102": "8211920000",   # ножи для пиццы
    "AAND1BL": "4016999708",      # игрушки для собак, резина
    "AAND1ST": "4016999708",
    "AAND1RP": "5609000000",      # канаты, хлопковая нить
    "ACRF1KP": "",                # капы — в матрице владелицы кода нет
}


# Колонки, которые показываем даже пустыми: это места под работу, которая ещё
# не сделана. Рич-контент владелица собирает в кабинете, сценариев в репозитории
# пока нет — но столбец должен быть виден, иначе про него забудут.
# Справочные колонки показываем всегда: ноль вопросов по группе — это факт,
# а исчезнувшая колонка выглядит как «не собирали».
FORCE_COLS = {"Рич-контент", "Вес товара с упаковкой (г)",
              "Заказов за год", "Вопросов покупателей",
              "Остаток на складах", "Себестоимость, ₽"}

# Чем заполняем такую колонку, если значения нет. Пустая ячейка по контракту листа
# значит «поле не разобрано», а здесь мы знаем причину — и она разная: брутто никто
# не взвешивал, а рич-контент просто не написан. Владелица 03.09 попросила, чтобы
# пустых ячеек в книге не оставалось вовсе.
FORCE_DEFAULTS = {
    "Вес товара с упаковкой (г)": MEASURE,   # остаётся для карточек вне каталога
    "Рич-контент": "сценария нет",
}

# Карточки, которых нет в каталоге кабинета: справочных цифр по ним взять неоткуда,
# и это надо писать словами, а не оставлять пустоту.
NOT_IN_CATALOG = "карточки нет в каталоге"


def tnved(article):
    """Код ТН ВЭД по артикулу. Пусто, если группа кода ещё не получила."""
    match = [k for k in TNVED if article.startswith(k)]
    return TNVED[max(match, key=len)] if match else ""


def sheet_snapshot(wb, name, cards):
    """Лист по группе без разбора: что стоит в кабинете сейчас.

    Отличается от обычного листа подписями колонок. В разобранных группах
    «Наименование» значит «внести это»; здесь — «сейчас стоит вот это»,
    и путать их нельзя, иначе владелица зальёт обратно то же самое.
    """
    rows = []
    for art in sorted(cards):
        c = cards[art]
        row = {"Артикул": art, "nm_id": c.get("nm_id"),
               "Наименование сейчас в кабинете": c.get("title") or "не заполнено",
               "Описание сейчас в кабинете": c.get("descr") or "не заполнено",
               "Статус": NO_ANALYSIS,
               "Код ТН ВЭД": tnved(art) or "нет в матрице кодов"}
        for k, v in (c.get("chars") or {}).items():
            row[k] = v.replace(" · ", "\n")
        fill_packaging(row)
        fill_facts(row)
        row.setdefault("Бренд", BRAND)
        row.setdefault("Страна производства", COUNTRY)
        rows.append(row)

    head = ["Артикул", "nm_id", "Статус", "Наименование сейчас в кабинете",
            "Описание сейчас в кабинете", "Код ТН ВЭД"]
    tail = ["Длина упаковки", "Ширина упаковки", "Высота упаковки",
            "Вес товара с упаковкой (г)", "Заказов за год", "Вопросов покупателей",
            "Остаток на складах", "Себестоимость, ₽", "Бренд", "Страна производства"]
    mid = sorted({k for r in rows for k in r} - set(head) - set(tail))
    cols = head + mid + tail

    ws = wb.create_sheet(name)
    ws.append(cols)
    for c in ws[1]:
        c.font = Font(bold=True)
    for r in rows:
        ws.append([r.get(c, "не заполнено") for c in cols])
    widths = {"Артикул": 15, "nm_id": 12, "Статус": 26,
              "Наименование сейчас в кабинете": 46, "Описание сейчас в кабинете": 90,
              "Код ТН ВЭД": 16}
    for i, c in enumerate(cols, 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = widths.get(c, 22)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"
    return len(rows)


# Группа по префиксу артикула — только для листа размеров, где строки идут
# по всему каталогу, а не по разобранным группам.
SIZE_GROUPS = [
    ("ACRA7TB", "Таблетницы"),
    ("ACRB1MS10", "Мячи и массажёры"),
    ("ACRH1MS", "Массажёр для пальцев"),
    ("AHMA", "Мешки для стирки"),
    ("AAND1BL", "Игрушки: мячи"),
    ("AAND1RP", "Игрушки: канаты"),
    ("ACRB1SK", "Скакалки"),
    ("ACRB3FR", "Фитнес-резинки"),
    ("ACRB4FR", "Фитнес-резинки"),
    ("ACRB5FR", "Фитнес-резинки"),
    ("ACRF1BN", "Бандажи косметические"),
    ("ACRF1KP", "Капы"),
    ("AKTA4ST", "Пробки для бутылок"),
    ("AKTA2KN101", "Шпатели кондитерские"),
    ("AKTA2KN102", "Ножи для пиццы"),
]


def size_group(art):
    for pref, name in SIZE_GROUPS:
        if art.startswith(pref):
            return name
    return ""


def size_issue(v):
    """Что не так с размерами конкретной карточки."""
    ih, iw, idp = v.get("item_hei"), v.get("item_wid"), v.get("item_dep")
    l, w, h = v.get("len"), v.get("wid"), v.get("hei")
    out = []
    if ih is None and iw is None:
        out.append("нет размеров предмета")
    if l is None:
        out.append("нет размеров упаковки")
    if None not in (ih, iw, idp, l, w, h):
        si = sorted((ih, iw, idp), reverse=True)
        sp = sorted((l, w, h), reverse=True)
        if si == sp:
            out.append("упаковка совпала с предметом — похоже, не измеряли")
        elif not all(a >= b for a, b in zip(sp, si)):
            out.append("упаковка меньше предмета хотя бы по одной стороне")
    net, br = v.get("net_g"), v.get("brutto_g")
    if net is None:
        out.append("нет веса нетто")
    if br is None:
        out.append("нет веса брутто")
    elif net is not None and br < net:
        # Брутто меньше нетто физически невозможно: либо нетто завышен,
        # либо брутто занижен. Разница мелкая (10–30 г), но в поставке
        # она умножается на тираж.
        out.append(f"брутто {br} г меньше нетто {net} г")
    if not out:
        out.append("всё заполнено")
    return "; ".join(out)


def sheet_sizes(wb):
    """Лист размеров и весов: предмет и упаковка рядом, с пометкой о расхождениях.

    Отдельным листом, потому что править это удобнее по всему каталогу разом,
    а не переключаясь между группами. Строки идут и по тем артикулам, которых
    в книге нет: подгруппа мячей 103 и таблетница 401 выведены из продажи,
    но размеры у них те же и чинить их придётся заодно.
    """
    if not PACKAGING:
        return 0
    ws = wb.create_sheet("Размеры и вес")
    ws.append(["Артикул", "Группа",
               "Предмет: высота", "ширина", "глубина", "вес нетто, г",
               "Упаковка: длина", "ширина", "высота", "вес брутто, г",
               "Что не так"])
    for c in ws[1]:
        c.font = Font(bold=True)
    def num(x):
        # Пусто в этом листе всегда значит одно и то же — величину не мерили.
        # Пишем это словами, иначе колонка брутто выглядит как забытая, а не как
        # незаполненная по всему каталогу.
        return MEASURE if x is None else x

    for art in sorted(PACKAGING):
        v = PACKAGING[art]
        ws.append([art, size_group(art) or "прочее",
                   num(v.get("item_hei")), num(v.get("item_wid")), num(v.get("item_dep")),
                   num(v.get("net_g")),
                   num(v.get("len")), num(v.get("wid")), num(v.get("hei")),
                   num(v.get("brutto_g")),
                   size_issue(v)])
    for col, wd in zip("ABCDEFGHIJK", (15, 21, 15, 9, 9, 13, 16, 9, 9, 14, 62)):
        ws.column_dimensions[col].width = wd
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "C2"
    return ws.max_row - 1


def sheet(wb, name, rows):
    ws = wb.create_sheet(name)
    for r in rows:
        r.setdefault("Код ТН ВЭД", tnved(r.get("Артикул", "")))
        # Поля, одинаковые у всех и известные без разбора. «Полное наименование
        # товара» WB держит отдельно от заголовка карточки, и оно пустует у семи
        # таблетниц из десяти и шести массажёров из двадцати трёх — при том что
        # значение всегда равно наименованию. «Страна производства» стоит всего
        # у двух карточек из сорока пяти, причём у одной ошибочно: ACRB1MS106BС
        # уехал в «Того» вместо Китая (проверка 02.09.2026).
        if r.get("Наименование"):
            r.setdefault("Полное наименование товара", r["Наименование"])
        r.setdefault("Страна производства", COUNTRY)
        fill_packaging(r)
        fill_facts(r)
        r.setdefault("Бренд", BRAND)
    # Известные колонки в заданном порядке, затем всё остальное разобранное.
    extra = sorted({k for r in rows for k in r} - set(COLS))
    cols = [c for c in COLS + extra if any(r.get(c) for r in rows) or c in FORCE_COLS]
    ws.append(cols)
    for c in ws[1]:
        c.font = Font(bold=True)
    facts_cols = {c for c, _ in FACT_FIELDS}
    for r in rows:
        line = []
        for c in cols:
            v = r.get(c, "")
            if v == "" and c in FORCE_DEFAULTS:
                v = FORCE_DEFAULTS[c]
            # Карточки нет в каталоге — тогда пусты и справочные цифры, и габариты.
            elif v == "" and r.get("Артикул") not in FACTS and (
                    c in facts_cols or c in {col for col, _ in PACK_FIELDS}):
                v = NOT_IN_CATALOG
            line.append(v)
        ws.append(line)
    widths = {"Артикул": 15, "nm_id": 12, "Наименование": 46,
              "Полное наименование товара": 46, "Описание": 90,
              "Рич-контент": 60, "Инфографика": 70, "Код ТН ВЭД": 14}
    for i, c in enumerate(cols, 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = widths.get(c, 22)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"
    return len(rows)


# ─────────────────────────────────────────────────────────────────────────────
# Лист по структуре кабинета
#
# Владелица 03.09: «составляй наш файл на основе этих данных, прям точь-в-точь —
# в каждой группе свои поля, их берём полностью». Поэтому колонки листа больше
# не наши: это колонки выгрузки кабинета, в исходном порядке, свои у каждой
# категории. Значения — тоже из кабинета; поверх ложатся решения разборов.
#
# Ключевое ограничение: файл должен оставаться пригодным для загрузки обратно.
# Значит в ячейку нельзя писать «нужен замер» или «верно, не трогать» — такой
# текст уедет в карточку. Поэтому статусы разбора выражены не текстом, а цветом,
# а в ячейке всегда лежит значение, которое должно там стоять.
EXPORT = json.loads((ROOT / "docs" / "seo" / "cabinet-export.json").read_text(encoding="utf-8")) \
    if (ROOT / "docs" / "seo" / "cabinet-export.json").exists() else {}

# Снятое с производства и выведенное из продажи. Карточка может ещё висеть
# в кабинете и попадать в выгрузку — в книгу она не идёт. За сессию 02–03.09
# так ушли четыре товара, каждый раз уже после того, как по нему что-то собрали,
# поэтому список общий и лежит рядом с данными, а не в коде.
DISCONTINUED = {k: v for k, v in (json.loads(
    (ROOT / "docs" / "seo" / "discontinued.json").read_text(encoding="utf-8")).items()
    if (ROOT / "docs" / "seo" / "discontinued.json").exists() else {}) if not k.startswith("_")}

FILL_NEW = PatternFill("solid", fgColor="D6EFD8")   # значение из разбора — менять
FILL_ASK = PatternFill("solid", fgColor="FFF2CC")   # ждём замер или ответ владелицы
FILL_REF = PatternFill("solid", fgColor="EDEDED")   # справочная колонка, не заливается

# Имена полей у нас и в кабинете совпадают почти везде. Здесь только исключения.
FIELD_MAP = {
    "Вес товара с упаковкой (г)": "Вес с упаковкой (кг)",
}

# Значения-статусы: в файл они не попадают, но задают цвет ячейки.
STATUS_ASK = (MEASURE, "нужен ответ владелицы", "нужна фотография цвета",
              "сценария нет", "карточки нет в каталоге")


def analysed_by_article():
    """Все решения разборов, собранные по артикулу: {артикул: {поле: значение}}."""
    out = {}
    for _, fname, parser in GROUPS:
        path = DESCR / fname
        if not path.exists():
            continue
        for row in parser(path.read_text(encoding="utf-8")):
            art = row.get("Артикул")
            if art:
                out.setdefault(art, {}).update(
                    {k: v for k, v in row.items() if k != "Артикул"})
    return out


# Пояснение при значении: «75 — окружность по слайду», «АРОЛС — стоит».
# В разборе так писать удобно — видно и значение, и почему оно такое. В ячейку
# должно попасть только само значение: файл грузится в кабинет.
CONFIRM_NOTES = {"стоит", "уже стоит", "стоит верно", "верно", "верно, не трогать"}


def strip_note(value):
    """Отрезать пояснение после « — », если это пояснение, а не часть значения.

    Ловушка здесь — комплектация: «Ручки 17 см — 2 шт.» тоже содержит тире,
    и резать её нельзя. Отличаем по правой части: у комплектации это короткий
    хвост в одно-два слова («1 шт.»), у пояснения — фраза. Перечисления через
    точку с запятой не трогаем вовсе.
    """
    if ";" in value or value.count(" — ") != 1:
        return value
    left, note = value.split(" — ", 1)
    if note.strip().lower().rstrip(".") in CONFIRM_NOTES:
        return left.strip()
    return left.strip() if len(note.split()) >= 3 else value


def sep(value):
    """Наши разделители значений — в кабинетный «;».

    В разборах значения разделены переводом строки или « · », в файле кабинета —
    точкой с запятой. Файл должен грузиться обратно, поэтому приводим к нему.
    """
    return value.replace("\n", ";").replace(" · ", ";").strip()


# Формулировки разбора, которые означают «поле должно остаться пустым».
# В ячейку они попасть не могут: файл грузится в кабинет, и такой текст уехал бы
# прямо в карточку.
CLEAR_PREFIXES = ("очистить", "не заполнено")


def cell_value(current, decided):
    """Что положить в ячейку и каким цветом её залить.

    current — то, что стоит в кабинете сейчас; decided — что сказал разбор.
    Возвращает (значение, заливка или None).
    """
    if not decided:
        return current, None
    low = decided.lower()
    if decided == OK:
        return current, None            # проверено, менять нечего
    if decided in STATUS_ASK:
        return current, FILL_ASK        # значение оставляем, но оно под вопросом
    if low.startswith(CLEAR_PREFIXES):
        # «очистить» — поле надо опустошить; «не заполнено» — оно и так пусто.
        return "", (FILL_NEW if current else None)
    if decided.endswith(f"— {OK}"):
        return sep(decided[:-len(f"— {OK}")]), None
    # Статус мог прийти с пояснением: «нужен замер — в поле 6 см, на слайде 50 мм».
    # Значение тогда не наше, а кабинетное, и ячейка просто помечается жёлтым.
    head = decided.split(" — ", 1)[0].strip()
    if head in STATUS_ASK or head.lower().startswith(
            ("нужен", "требует", "привести", "сверить", "подтвердить")):
        return current, FILL_ASK
    if head.lower().startswith(CLEAR_PREFIXES):
        return "", (FILL_NEW if current else None)
    value = sep(strip_note(decided))
    return (current, None) if value == current else (value, FILL_NEW)


def sheet_export(wb, group, cards, decided_all):
    """Лист группы: колонки и значения кабинета, поверх — решения разборов."""
    cols = list(cards["cols"])
    hints = cards.get("hints") or [""] * len(cols)
    # Справочные колонки в кабинет не заливаются, поэтому идут последними
    # и залиты серым — чтобы при копировании в шаблон их было видно и отрезать.
    refs = [c for c, _ in FACT_FIELDS]

    ws = wb.create_sheet(group[:31])
    ws.append(cols + refs)
    ws.append(hints + ["в кабинет не заливается"] * len(refs))
    for c in ws[1]:
        c.font = Font(bold=True)
    for c in ws[2]:
        c.font = Font(size=8, italic=True, color="808080")

    for art in sorted(cards["rows"]):
        if art in DISCONTINUED:
            continue
        current = dict(zip(cols, cards["rows"][art]))
        decided = decided_all.get(art, {})
        line, fills = [], []
        for c in cols:
            ours = decided.get(c)
            if ours is None:
                # поле у нас называется иначе — ищем по карте
                for mine, theirs in FIELD_MAP.items():
                    if theirs == c and mine in decided:
                        ours = decided[mine]
                        break
            value, fill = cell_value(current.get(c, ""), ours)
            line.append(value)
            fills.append(fill)
        facts = FACTS.get(art, {})
        for _, key in FACT_FIELDS:
            v = facts.get(key)
            line.append("" if v is None else v)
            fills.append(FILL_REF)
        ws.append(line)
        for i, fill in enumerate(fills, 1):
            if fill:
                ws.cell(ws.max_row, i).fill = fill

    widths = {"Артикул продавца": 15, "Артикул WB": 12, "Наименование": 46,
              "Описание": 90, "Фото": 30, "Комплектация": 40,
              "Полное наименование товара": 40, "Категория продавца": 22}
    for i, c in enumerate(cols + refs, 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = widths.get(c, 20)
    for row in ws.iter_rows(min_row=3):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "C3"
    return ws.max_row - 2          # минус строка заголовков и строка подсказок


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--out", default="kartochki_arols.xlsx")
    a = ap.parse_args()

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    intro = wb.create_sheet("Как заливать")
    for line in [
        ["Структура", "Один лист на категорию кабинета. Колонки и их порядок взяты "
                      "из вашей выгрузки от 03.09 — у каждой группы свои поля. "
                      "Вторая строка листа — подсказки WB из того же файла."],
        ["Значения", "В ячейке всегда лежит то, что должно стоять в карточке после правки. "
                     "Служебных пометок вроде «нужен замер» в ячейках нет: файл пригоден "
                     "для загрузки обратно, и такой текст уехал бы прямо в карточку."],
        ["🟩 Зелёная ячейка", "Значение изменено по разбору — это и есть правка. "
                              "Обоснование каждой — в docs/seo/descriptions/."],
        ["🟨 Жёлтая ячейка", "Значение оставлено как в кабинете, но оно под вопросом: "
                             "ждём замер, фотографию или ваш ответ. Список вопросов — "
                             "в конце файла разбора соответствующей группы."],
        ["⬜ Серая ячейка", "Справочные колонки в конце листа — заказы, вопросы, остаток, "
                            "себестоимость. В кабинет НЕ заливаются, показывают, насколько "
                            "карточка важна. При копировании в шаблон их надо отрезать."],
        ["Белая ячейка", "Значение из кабинета, разбор его не трогал."],
        ["Пустая ячейка", "В кабинете пусто, и чем заполнить — мы пока не знаем. "
                          "Писать туда пометку нельзя: она попадёт в карточку."],
        ["Разделитель", "Несколько значений одного поля разделены точкой с запятой — "
                        "так же, как в выгрузке кабинета."],
        ["Порядок", "Сначала характеристики, затем наименование, затем описание."],
        ["Мячи", "Заливать описания только после правки инфографики: "
                 "docs/seo/infographics/myachi-106-107-109.md."],
        ["Инфографика", "Отдельный лист: что переделать на слайдах, по группам. Мячи и массажёры — "
                        "срочно, там медицинские заявления. Таблетницы — не срочно."],
        ["Размеры и вес", "Отдельный лист по всему каталогу: габариты предмета и упаковки, вес "
                          "нетто и брутто рядом, с колонкой «Что не так»."],
        ["Код ТН ВЭД", "Из вашей матрицы «ТН ВЭД × ОКПД 2» — тот код, что помечен зелёным. "
                       "Разбор и расхождения с прежними кодами — docs/marking/codes-marking.md."],
        ["Снятое с производства", "В книгу не идёт, даже если карточка ещё висит в кабинете "
                                  "и попадает в выгрузку. Сейчас это палочка-кость AAND1ST202BL, "
                                  "скакалка со счётчиком ACRB1SK201BC, шпатель голубой AKTA2KN101BL "
                                  "и латексный набор ACRB5FR400CL. Список — docs/seo/discontinued.json."],
        ["Бандажи, ножи,\nпробки", "Разбора по ним ещё не было — ждём срез Эвирмы. В ячейках лежит "
                                   "то, что стоит в кабинете, цветных пометок нет."],
    ]:
        intro.append(line)
    intro.column_dimensions["A"].width = 18
    intro.column_dimensions["B"].width = 95
    for row in intro.iter_rows():
        row[0].font = Font(bold=True)
        row[1].alignment = Alignment(wrap_text=True, vertical="top")

    per_article = all_per_article()
    n_info = sheet_info(wb)
    if n_info:
        print(f"Инфографика: {n_info} пунктов")
    n_sizes = sheet_sizes(wb)
    if n_sizes:
        print(f"Размеры и вес: {n_sizes} артикулов")

    # Один лист на категорию кабинета, с её собственными полями. Решения разборов
    # ложатся поверх значений кабинета и подсвечиваются цветом — см. sheet_export.
    decided = analysed_by_article()
    total = 0
    for group in sorted(EXPORT):
        n = sheet_export(wb, group, EXPORT[group], decided)
        total += n
        live = [a for a in EXPORT[group]["rows"] if a not in DISCONTINUED]
        touched = sum(1 for a in live if a in decided)
        gone = len(EXPORT[group]["rows"]) - len(live)
        note = f", снятых пропущено {gone}" if gone else ""
        print(f"{group}: {n} карточек, разобрано {touched}{note}")

    # Группы, которых в выгрузке нет: их данные лежат в снимке от 03.09.
    # Ключ «_отложено» перечисляет то, что владелица просила не трогать.
    postponed = SNAPSHOT.get("_отложено", {})
    for name in SNAPSHOT:
        if name == "_отложено" or name in postponed or name in EXPORT:
            continue
        n = sheet_snapshot(wb, name, SNAPSHOT[name])
        total += n
        print(f"{name}: {n} карточек — снимок кабинета, выгрузки по группе нет")

    wb.save(a.out)
    print(f"записано {total} карточек → {a.out}")


if __name__ == "__main__":
    main()
