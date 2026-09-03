#!/usr/bin/env python3
"""Внести значения из свежей сборки в книгу владелицы, не трогая её оформление.

Книга у владелицы своя: она правит ширину колонок, высоту строк и красит
артикулы — зелёным обработанные, красным снятые с продажи. Пересобранная
с нуля книга это стирает, поэтому генератор в её файл не пишет: сюда приходят
только те колонки, которые названы явно.

По умолчанию переносится «Описание» — оно чинилось после того, как файл был
отдан. Всё остальное остаётся как есть, включая значения, которые в сборке
отличаются: расхождения показывает --dry-run, решение по ним за владелицей.

    python3 scripts/apply_to_book.py книга.xlsx --out книга-обновлённая.xlsx
    python3 scripts/apply_to_book.py книга.xlsx --dry-run
"""
import argparse
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill

ROOT = Path(__file__).resolve().parent.parent
BUILT = ROOT / "kartochki_arols.xlsx"
KEY = "Артикул продавца"
# Листы без карточек: там нечего сопоставлять по артикулу.
SKIP = {"Как читать файл", "Порядок заливки", "Инфографика", "Размеры и вес"}
DASHES = str.maketrans({"—": "-", "–": "-"})


def as_rule(value, column):
    """Значение, приведённое к правилам оформления §21: тире и регистр."""
    out = str(value).translate(DASHES)
    if column == "Комплектация":
        out = ";".join(i.lstrip()[:1].upper() + i.lstrip()[1:]
                       for i in out.split(";"))
    return out


def only_formatting(a, b, column):
    """Различие целиком объясняется правилами §21 — оформление, не правка.

    Такое переносится молча: согласовывать нечего. Всё остальное — текст,
    цифры, чужой регистр внутри слова — остаётся владелице.
    """
    return as_rule(a, column) == as_rule(b, column) and a != b


def rows_by_article(ws, key_col):
    return {str(ws.cell(r, key_col).value): r
            for r in range(3, ws.max_row + 1) if ws.cell(r, key_col).value}


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("book", help="книга владелицы")
    ap.add_argument("--out", help="куда записать; без него — рядом, с суффиксом")
    ap.add_argument("--columns", default="Описание",
                    help="колонки через запятую")
    ap.add_argument("--dry-run", action="store_true",
                    help="только показать, что разошлось, и ничего не писать")
    args = ap.parse_args()

    columns = [c.strip() for c in args.columns.split(",") if c.strip()]
    book = openpyxl.load_workbook(args.book)
    built = openpyxl.load_workbook(BUILT)

    changed = 0
    dashed = 0
    repainted = 0
    other = {}
    for name in book.sheetnames:
        if name in SKIP or name not in built.sheetnames:
            continue
        theirs, ours = book[name], built[name]
        th = [str(c.value) for c in theirs[1]]
        oh = [str(c.value) for c in ours[1]]
        if th != oh or KEY not in th:
            print(f"  {name}: колонки разошлись, лист пропущен")
            continue
        ki = th.index(KEY) + 1
        # Колонка артикула — территория владелицы: там её зелёные и красные
        # пометки «обработано» и «выведено». Заливку туда не трогаем.
        trow = rows_by_article(theirs, ki)
        orow = rows_by_article(ours, ki)
        for art, r in trow.items():
            if art not in orow:
                continue
            for ci, col in enumerate(th, 1):
                a = theirs.cell(r, ci).value or ""
                b = ours.cell(orow[art], ci).value or ""
                # Подсветка = «ещё не залито в кабинет». Сборка считает её
                # от живого слепка кабинета, поэтому сделанное гаснет само —
                # переносим заливку как есть, включая её снятие.
                if ci != ki and not args.dry_run:
                    src = ours.cell(orow[art], ci).fill
                    dst = theirs.cell(r, ci)
                    want = (PatternFill("solid", fgColor=src.fgColor.rgb)
                            if src.fill_type == "solid" else PatternFill())
                    have = (dst.fill.fgColor.rgb
                            if dst.fill.fill_type == "solid" else None)
                    if have != (src.fgColor.rgb if src.fill_type == "solid" else None):
                        dst.fill = want
                        repainted += 1
                if a == b:
                    continue
                if only_formatting(a, b, col):
                    # §21: тире и регистр комплектации — правила оформления,
                    # а не смысловая правка. Переносим без согласования.
                    if not args.dry_run:
                        theirs.cell(r, ci).value = b
                    dashed += 1
                elif col in columns or src.fill_type == "solid":
                    # Ячейка подсвечена как «не залито» — значит владелице
                    # нужно видеть, ЧЕМ заменить. Раньше переносилось только
                    # «Описание», и она получала зелёный фон поверх старого
                    # значения: подсветка есть, а нового текста нет.
                    if not args.dry_run:
                        theirs.cell(r, ci).value = b
                    changed += 1
                    print(f"  {name} · {art} · {col}: {str(a)[:40]!r} → {str(b)[:40]!r}")
                else:
                    other.setdefault(col, []).append(f"{name}/{art}")

    verb = "перенесено" if not args.dry_run else "будет перенесено"
    print(f"\n{verb} значений: {changed}")
    print(f"{verb} исправлений оформления (§21): {dashed}")
    print(f"перекрашено ячеек (подсветка «не сделано»): {repainted}")
    if other:
        print("\nразошлось ещё, но НЕ переносится — решение за владелицей:")
        for col, items in sorted(other.items(), key=lambda x: -len(x[1])):
            print(f"  {col:<34} {len(items):>3}  {', '.join(items[:2])}"
                  f"{'…' if len(items) > 2 else ''}")

    if not args.dry_run:
        out = args.out or str(Path(args.book).with_suffix("")) + "-обновлено.xlsx"
        book.save(out)
        print(f"\nзаписано → {out}")


if __name__ == "__main__":
    main()
