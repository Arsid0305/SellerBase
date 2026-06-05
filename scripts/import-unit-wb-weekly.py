#!/usr/bin/env python3
"""Парсит UNIT_WB по неделям 2026.xlsx, генерит SQL UPSERT в sku_weekly_metrics."""
import openpyxl
import os

f = "/root/.claude/uploads/071b7af7-97ac-4d25-aae7-6a11641c664c/fb9bcd6c-UNIT_WB______2026.xlsx"
wb = openpyxl.load_workbook(f, data_only=True)

# Колонки (по нумерации row 3 в файле):
# 1=wb_article, 2=арт. пост., 3=название, 4=barcode
# 5=Остаток, 6=Себ/ед, 7=Себ остат, 8=Оборачиваемость, 9=%выкупа, 10=Скорость
# 11=Продажи шт, 12=Возвраты шт, 13=Продажи-возвраты
# 14=Выручка ЦП, 15=Выручка СПП, 16=Выручка возвраты, 17=Выручка цена ВБ, 18=Выручка ВБ
# 19=Доля выручки, 20=Себестоимость продаж, 21=Комиссия руб, 22=Комиссия %, 23=Логистика, 24=Хранение
# 27=Чистая прибыль

values = []
for week_num in range(1, 54):
    sheet_name = str(week_num)
    if sheet_name not in wb.sheetnames:
        continue
    ws = wb[sheet_name]
    for r in range(4, ws.max_row + 1):
        wb_art = ws.cell(r, 1).value
        if not wb_art or not isinstance(wb_art, (int, float)):
            continue
        barcode = ws.cell(r, 4).value

        def num(col):
            v = ws.cell(r, col).value
            return float(v) if isinstance(v, (int, float)) else None

        def int_or(col):
            v = ws.cell(r, col).value
            return int(v) if isinstance(v, (int, float)) else None

        values.append({
            'wb_article': int(wb_art),
            'barcode': str(barcode) if barcode else None,
            'year': 2026,
            'week_num': week_num,
            'stock_start': int_or(5),
            'cost_per_unit': num(6),
            'cost_stock_total': num(7),
            'turnover_days': num(8),
            'buyout_pct': num(9),
            'sales_velocity': num(10),
            'units_sold': int_or(11),
            'units_returned': int_or(12),
            'units_net': int_or(13),
            'revenue_wb': num(18),
            'cost_sold_total': num(20),
            'commission_rub': num(21),
            'commission_pct': num(22),
            'logistics_rub': num(23),
            'storage_rub': num(24),
            'net_profit': num(27),
        })


def f_sql(x):
    if x is None:
        return 'NULL'
    if isinstance(x, str):
        return "'" + x.replace("'", "''") + "'"
    return str(x)


out_path = os.path.join(os.path.dirname(__file__), 'import-unit-wb-weekly.sql')
with open(out_path, 'w') as out:
    out.write("-- Auto-generated from UNIT_WB по неделям 2026.xlsx\n")
    out.write("-- Запускать после применения миграции sku_weekly_metrics\n\n")
    if not values:
        out.write("-- (нет данных)\n")
    else:
        out.write(
            "WITH input(wb_article, barcode, year, week_num, stock_start, cost_per_unit, "
            "cost_stock_total, turnover_days, buyout_pct, sales_velocity, units_sold, "
            "units_returned, units_net, revenue_wb, cost_sold_total, commission_rub, "
            "commission_pct, logistics_rub, storage_rub, net_profit) AS (VALUES\n"
        )
        rows = []
        for v in values:
            rows.append(
                f"({f_sql(v['wb_article'])},{f_sql(v['barcode'])},{v['year']},{v['week_num']},"
                f"{f_sql(v['stock_start'])},{f_sql(v['cost_per_unit'])},{f_sql(v['cost_stock_total'])},"
                f"{f_sql(v['turnover_days'])},{f_sql(v['buyout_pct'])},{f_sql(v['sales_velocity'])},"
                f"{f_sql(v['units_sold'])},{f_sql(v['units_returned'])},{f_sql(v['units_net'])},"
                f"{f_sql(v['revenue_wb'])},{f_sql(v['cost_sold_total'])},{f_sql(v['commission_rub'])},"
                f"{f_sql(v['commission_pct'])},{f_sql(v['logistics_rub'])},{f_sql(v['storage_rub'])},"
                f"{f_sql(v['net_profit'])})"
            )
        out.write(',\n'.join(rows))
        out.write("""
)
INSERT INTO sku_weekly_metrics (sku_id, wb_article, barcode, year, week_num, stock_start, cost_per_unit, cost_stock_total, turnover_days, buyout_pct, sales_velocity, units_sold, units_returned, units_net, revenue_wb, cost_sold_total, commission_rub, commission_pct, logistics_rub, storage_rub, net_profit)
SELECT c.id, i.wb_article, i.barcode, i.year, i.week_num, i.stock_start, i.cost_per_unit, i.cost_stock_total, i.turnover_days, i.buyout_pct, i.sales_velocity, i.units_sold, i.units_returned, i.units_net, i.revenue_wb, i.cost_sold_total, i.commission_rub, i.commission_pct, i.logistics_rub, i.storage_rub, i.net_profit
FROM input i
LEFT JOIN sku_catalog c ON c.wb_article = i.wb_article
ON CONFLICT (sku_id, year, week_num) DO UPDATE SET
  stock_start = EXCLUDED.stock_start,
  cost_per_unit = EXCLUDED.cost_per_unit,
  cost_stock_total = EXCLUDED.cost_stock_total,
  turnover_days = EXCLUDED.turnover_days,
  buyout_pct = EXCLUDED.buyout_pct,
  sales_velocity = EXCLUDED.sales_velocity,
  units_sold = EXCLUDED.units_sold,
  units_returned = EXCLUDED.units_returned,
  units_net = EXCLUDED.units_net,
  revenue_wb = EXCLUDED.revenue_wb,
  cost_sold_total = EXCLUDED.cost_sold_total,
  commission_rub = EXCLUDED.commission_rub,
  commission_pct = EXCLUDED.commission_pct,
  logistics_rub = EXCLUDED.logistics_rub,
  storage_rub = EXCLUDED.storage_rub,
  net_profit = EXCLUDED.net_profit;
""")
print(f"Сгенерировано {len(values)} строк")
